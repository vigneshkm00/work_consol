import os
import openpyxl as op
os.chdir("C:\\Users\\vikm\\Desktop\\New folder")
path = os.getcwd()
all_files = os.listdir(path)
work_files = [f for f in all_files if f[-4:] == 'xlsx']
print(work_files)
# dataframe = pd.DataFrame()
wb2 = op.Workbook()
sheet2 = wb2.active
for file_name in work_files:
    valu = ()
    wb1 = op.load_workbook(file_name)
    sheet1 = wb1.active
    mx_row = sheet1.max_row
    print(mx_row)
    for x in range(1, mx_row+1):
        valu = (sheet1.cell(row=x, column=1).value, sheet1.cell(row=x, column=2).value, sheet1.cell(row=x, column=4).value,
                sheet1.cell(row=x, column=11).value, sheet1.cell(row=x, column=12).value, sheet1.cell(row=x, column=41).value)
        # valu = (sheet1.cell(row=x, column=1).value, sheet1.cell(row=x, column=2).value, sheet1.cell(row=x, column=4).value,
        #         sheet1.cell(row=x, column=10).value, sheet1.cell(row=x, column=11).value, sheet1.cell(row=x, column=40).value)
        sheet2.append(valu)
file_nametosav = input("Enter file name to save:")
wb2.save(file_nametosav+".xlsx")
print("Success...")
